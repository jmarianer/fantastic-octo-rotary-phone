#This program will read an array of values from a spreadsheet,
#then prompt a user to input a number from 1-5 and print the
#correct value.

#import openpyxl module function load_workbook
from openpyxl import load_workbook

#Define variable name for the workbook
wb = load_workbook('samplebook.xlsx')

#Define variable name for worksheet
sheet = wb['Sheet1']

#Read in values of cells C1 through C5 as list (array).
Cells_List=[]
for i in range(1,6):
    Cells_List.append(int(sheet.cell(i,3).value))
print(Cells_List)

#Prompt user for the array item.
#Cell_Wanted=input("Please enter the cell you want the value for: ")

#Print the desired cell.
#print(int(Cells_List[Cell_Wanted-1]))

#Square the value of each cell and write it to a new list.
Cell_Squares=[]
for i in range (1,6):
    Cell_Squares.append((Cells_List[i-1])*(Cells_List[i-1]))
print(Cell_Squares)

#Cube the value of each cell and write it to a new list.
Cell_Cubes=[]
for i in range (1,6):
    Cell_Cubes.append((Cells_List[i-1])*(Cells_List[i-1])*(Cells_List[i-1]))
print(Cell_Cubes)

#Take the even values only from Cells_List and write them to a new list.
Cell_Evens=[]
for i in range (1,6):
    if (Cells_List[i-1])%2 == 0:
        Cell_Evens.append(Cells_List[i-1])
print(Cell_Evens)

#Take the even values from Cell_Evens, square them, and write them to a new list.
Cell_Even_Squares=[]
for i in range (1,len(Cell_Evens)+1):
    Cell_Even_Squares.append(Cell_Evens[i-1]*Cell_Evens[i-1])
print(Cell_Even_Squares)

#Print the values of list Cells_List.
#for j in range(1,6):
#    print(Cells_List[j-1])