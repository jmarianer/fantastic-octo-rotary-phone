#This program will read in an Excel file and print the contents of a
#specific cell and the total of a group of cells summed together.

#import openpyxl module function load_workbook
from openpyxl import load_workbook

#Define variable name for the workbook
wb = load_workbook('samplebook.xlsx')

#Define variable name for worksheet
sheet = wb['Sheet1']

#Define variable for the cell to read
cell = sheet.cell(1,1)

#Define varible for cell value to print
my_val = cell.value

#Print the cell value
print(my_val)

#Read in values of cells C1 through C5 as list (array).
Cells_List=[]
for i in range(1,6):
    Cells_List.append(sheet.cell(i,3))

#Print the values of list Cells_List.
for j in range(1,6):
    print(Cells_List[j-1].value)

#Read in values of cells C1 through C5 and print the total value.
#Define the first value to be used as the base of the for loop.
Total_Cell_Values = 0

for i in range(1,6):
    C_Cell_value = sheet.cell(i,3).value
    Total_Cell_Values = Total_Cell_Values + C_Cell_value

#Write Total_Cell_Values to Excel file cell D1.
sheet['D1'] = Total_Cell_Values

wb.save("sample_output.xlsx")
