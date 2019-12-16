# -*- coding: utf-8 -*-
"""
Created on Thu Dec 12 15:43:17 2019
Created on Thu Dec 12 13:55:04 2019
This code will read the folder directory from an Excel spreadsheet and delete 
the folder, if empty.
@author: ad414d
"""

import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

#Define variable for the TreeSize workbook.
wb=load_workbook('C:\\aaTanker\TreeSize\jrl me folder - new.xlsx')

#Define variable for the worksheet.
sheet = wb['Custom Search']

for i in range(5, sheet.max_row):

    Cell_To_Check = sheet.cell(row=i,column=3)

    #Check if the file exists, then delete it.
    if os.path.exists(Cell_To_Check.value):
        #print("Path exists.")
        #print(sheet.cell(row=685,column=3).value+sheet.cell(row=685,column=2).value)
        try:
            os.rmdir(Cell_To_Check.value)
        except OSError:
            is_empty = False
        if is_empty:
            Color_Fill_Duplicate_Deleted = PatternFill(fgColor='D8E4BC',
                                                       fill_type='solid')
            Cell_To_Check.fill = Color_Fill_Duplicate_Deleted
        
wb.save('C:\\aaTanker\TreeSize\jrl me folder - new.xlsx')