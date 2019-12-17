# -*- coding: utf-8 -*-
"""
Created on Thu Dec 12 13:55:04 2019
This code will read the file directory and name from an Excel spreadsheet
and delete the file, provided the item is a duplicate and still exists.
@author: ad414d
"""

import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

#Define variable for the TreeSize workbook.
wb=load_workbook(r'C:\aaTanker\TreeSize\jrl me folder - new.xlsx')

#Define variable for the worksheet.
sheet = wb['Custom Search']

for i in range(714, sheet.max_row):

    Cell_File_Name = sheet.cell(row=i,column=2)
    Cell_Path_Name = sheet.cell(row=i,column=3)
    Cell_Duplicate = sheet.cell(row=i,column=1)

    #Check if the file is a duplicate.  If it is, then continue to removal.
    if sheet.cell(row=i,column=1).value == "Duplicate":
        #Check if the file exists, then delete it.
        if os.path.exists(Cell_Path_Name.value+Cell_File_Name.value):
            os.remove(Cell_Path_Name.value+Cell_File_Name.value)
            Color_Fill_Duplicate_Deleted = PatternFill(fgColor='D8E4BC',
                                                       fill_type='solid')
            Cell_Duplicate.fill = Color_Fill_Duplicate_Deleted
    else:
        print("The file does not exist.")
        
wb.save(r'C:\aaTanker\TreeSize\jrl me folder - new.xlsx')