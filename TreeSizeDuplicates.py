# -*- coding: utf-8 -*-
"""
Created on Wed Dec 11 11:43:25 2019
This program will read a file in from TreeSize and determine what duplicates
exist and mark them "Duplicate" or "Different" in a new column.
Program inserts a new column before column A.
Comparison is of adjacent rows using columns B, D, E, and F.
Sort by column A prior to running is currently required.
@author: ad414d
"""

#Import function to read xlsx file.
from openpyxl import load_workbook

#Define variable for the TreeSize workbook.
wb=load_workbook(r'C:\aaTanker\TreeSize\jrl me folder - clean.xlsx')

#Define variable for the worksheet.
sheet = wb['Custom Search']

#Insert a new column A.
sheet.insert_cols(1)
sheet.column_dimensions['A'].width=10

#Declare a set of tuples.
seen_tuples = set()

#Compare a row to the next.
#NR = next row.
for j in range(5,sheet.max_row):
    Duplicate_or_Different = sheet.cell(row=j,column=1)

    NR2B=sheet.cell(row=j,column=2)
    NR2D=sheet.cell(row=j,column=4)
    NR2E=sheet.cell(row=j,column=5)
    NR2F=sheet.cell(row=j,column=6)

    Next_Row=(NR2B.value, NR2D.value, NR2E.value, NR2F.value)

    if Next_Row in seen_tuples:
        Duplicate_or_Different.value = "Duplicate"
    else:
        seen_tuples.add(Next_Row)
        Duplicate_or_Different.value = "Different"

wb.save(r'C:\aaTanker\TreeSize\jrl me folder - set - new.xlsx')