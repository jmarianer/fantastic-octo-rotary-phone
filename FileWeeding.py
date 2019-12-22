# -*- coding: utf-8 -*-
"""
Created on Mon Dec 16 17:12:20 2019
This code will compare file names and, if identical, will compare the contents
to see if they are duplicates or not.  If they are duplicates, it will put
the word "duplicate" in a new column A.  Otherwise, it will mark the file
"Unique".
@author: ad414d
"""

#Import the functions required.
from openpyxl import load_workbook
import filecmp

#Declare global variables.
wb=load_workbook(r'C:\aaTanker\TreeSize\jrl me folder - new2.xlsx')
sheet=wb['Custom Search']

#Declare constants
DUPLICATION_COLUMN = 1
FILENAME_COLUMN = 2
FILE_FOLDER_PATH = 3
FIRST_DATA_ROW = 5

#Insert a new column A.
sheet.insert_cols(DUPLICATION_COLUMN)
sheet.column_dimensions['A'].width=10

#Create a list variable for the file names.
File_Names_List = []

#Add the first value to the File_Names_List.
File_Names_List.append(sheet.cell(FIRST_DATA_ROW,DUPLICATION_COLUMN))
sheet.cell(FIRST_DATA_ROW,DUPLICATION_COLUMN).value = "Unique"

#Determine if the file name is a duplicate of a previous file name.
for i in range (FIRST_DATA_ROW + 1,sheet.max_row):

    #Declare variables.
    Duplicate_or_Unique = sheet.cell(i,DUPLICATION_COLUMN)
    File_Name_Cell = sheet.cell(i,FILENAME_COLUMN)
    File_Folder_Cell = sheet.cell(i,FILE_FOLDER_PATH)

    #Look to see if the File name is already in File_Names_List.  If it is,
    #then compare the files.
    if File_Name_Cell.value in File_Names_List:
         Duplicate_or_Unique.value = "Duplicate: " \
         + str(File_Names_List.index(File_Name_Cell.value)+FIRST_DATA_ROW)
         
         

    else:
        Duplicate_or_Unique.value = "Unique"

    #Generate the File_Names_List.
    File_Names_List.append(File_Name_Cell.value)
    
wb.save(r'C:\aaTanker\TreeSize\jrl me folder - test output.xlsx')    
