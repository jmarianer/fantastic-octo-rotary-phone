#This program will read in all of the states, their populations, and the total
#population of the United States.  It will then sum the population of all the
#states and determine if the value is the same as the total population.

#Import function to read xlsx file.
from openpyxl import load_workbook

#Define variable for the workbook.
wb = load_workbook('nst-est2018-01.xlsx')

#Define variable for the worksheet.
sheet = wb['NST01']

#Define variables.
Total_State_Population = 0

#Create a list of all the states and their populations.
States_List=[]
for i in range (10,61):
    States_List.append(sheet.cell(i,1).value)
print(States_List)

State_Populations_List=[]
for i in range (10,61):
    State_Populations_List.append(int(sheet.cell(i,2).value))
    #Sum up all of the populations.
    Total_State_Population = Total_State_Population + int(sheet.cell(i,2).value)
print(State_Populations_List)
print(Total_State_Population)

#Read in the total US population and compare to the sum total above.
Total_US_Population=sheet.cell(5,3)
print(Total_US_Population.value)

if Total_US_Population.value == Total_State_Population:
    print('checks out')
else:
    print('oops')